import subprocess
import os
import shutil
import tempfile
import re
from datetime import datetime, timedelta
from dateutil import tz
import openpyxl
from openpyxl.utils import get_column_letter
from viam.logging import getLogger

LOGGER = getLogger(__name__)

class WorkbookProcessor:
    def __init__(self, work_dir, export_script, api_key_id, api_key, org_id, 
                 timezone="America/New_York", export_start_time="7:00", export_end_time="19:00"):
        self.work_dir = work_dir
        self.export_script_path = export_script
        self.export_script_dir = os.path.dirname(export_script)
        self.api_key_id = api_key_id
        self.api_key = api_key
        self.org_id = org_id
        self.timezone = timezone
        self.export_start_time = export_start_time
        self.export_end_time = export_end_time
        # Check if LibreOffice is available for formula recalculation
        self.libreoffice_available = self._check_libreoffice()
        if self.libreoffice_available:
            LOGGER.info("LibreOffice is available for formula recalculation")
        else:
            LOGGER.warning("LibreOffice not found - complex formula recalculation may be limited")

        # Check for viam-python-data-export virtual environment
        self.venv_path = os.path.join(self.export_script_dir, ".venv")
        if not os.path.exists(self.venv_path):
            LOGGER.warning(f"viam-python-data-export virtual environment not found at {self.venv_path}, attempting to set it up")
            self._setup_venv()
        else:
            LOGGER.info(f"Found viam-python-data-export virtual environment at {self.venv_path}")
    
    def _setup_venv(self):
        """Set up the virtual environment for viam-python-data-export if it doesn't exist"""
        try:
            setup_script = os.path.join(self.export_script_dir, "setup.sh")
            if os.path.exists(setup_script):
                LOGGER.info(f"Running setup script for viam-python-data-export: {setup_script}")
                # Cannot source the script directly (can simulate it)
                # Running bash with -c and sourcing the script
                subprocess.run(
                    ["bash", "-c", f"cd {self.export_script_dir} && source ./setup.sh"],
                    check=True,
                    shell=False
                )
                LOGGER.info("viam-python-data-export setup script completed successfully!")
            else:
                LOGGER.error(f"Setup script not found at {setup_script}")
        except Exception as e:
            LOGGER.error(f"Failed to set up viam-python-data-export virtual environment: {e}")

    def _check_libreoffice(self):
        """Check if LibreOffice is available on the system."""
        try:
            result = subprocess.run(
                ["which", "libreoffice"], 
                capture_output=True, 
                text=True, 
                check=False
            )
            return result.returncode == 0
        except Exception:
            return False

    def _extract_date_from_filename(self, filename):
        """Extract date from template filename like 3895th_MMDDYY.xlsx."""
        try:
            # Extract MMDDYY portion from the filename
            match = re.search(r'3895th_(\d{6})\.xlsx', os.path.basename(filename))
            if match:
                date_str = match.group(1)
                # Parse as month, day, year
                month = int(date_str[0:2])
                day = int(date_str[2:4])
                # Convert 2-digit year to 4-digit year
                year = int(date_str[4:6])
                # Assume 20xx for the century (valid until 2099)
                year += 2000
                
                # Create date object
                return datetime(year, month, day, tzinfo=tz.gettz(self.timezone))
            else:
                LOGGER.warning(f"Could not extract date from filename: {filename}")
                return None
        except Exception as e:
            LOGGER.error(f"Error parsing date from filename {filename}: {e}")
            return None

    def get_yesterday_date(self):
        """Get yesterday's date in the configured timezone."""
        now = datetime.now(tz.gettz(self.timezone))
        return now - timedelta(days=1)

    def run_vde_export(self, output_file, target_date=None):
        """Run the vde.py script to export raw data for the specified date or yesterday."""
        # Use the provided date or default to yesterday
        if target_date is None:
            target_date = self.get_yesterday_date()
            LOGGER.info(f"No target date provided, using yesterday: {target_date.strftime('%Y-%m-%d')}")
        
        # Parse the time strings into hours and minutes
        start_hour, start_minute = map(int, self.export_start_time.split(':'))
        end_hour, end_minute = map(int, self.export_end_time.split(':'))
        
        # Create the datetime objects for start and end times
        start_time = target_date.replace(hour=start_hour, minute=start_minute, second=0, microsecond=0)
        end_time = target_date.replace(hour=end_hour, minute=end_minute, second=0, microsecond=0)

        LOGGER.info(f"Exporting data from {start_time} to {end_time} ({self.export_start_time} to {self.export_end_time})")

        # Construct the shell script to run vde.py with its virtual environment
        venv_python = os.path.join(self.venv_path, "bin", "python")
        export_script_path = self.export_script_path

        # Build the command        
        cmd = [
            venv_python,
            export_script_path,
            "-vv", 
            "excel",
            "--apiKeyId", self.api_key_id,
            "--apiKey", self.api_key,
            "--orgId", self.org_id,
            "--resourceName", "langer_fill",
            "--start", start_time.strftime("%Y-%m-%dT%H:%M:%S%z"),
            "--end", end_time.strftime("%Y-%m-%dT%H:%M:%S%z"),
            "--timezone", self.timezone,
            "--bucketPeriod", "PT5M",
            "--bucketMethod", "max",
            "--includeKeys", ".*_raw",
            "--output", output_file,
            "--tab", "RAW"
        ]

        # Create a masked version of the command for logging
        cmd_mask = cmd.copy()

        # Find the index of the sensitive parameters and mask them
        if "--apiKeyId" in cmd_mask:
            idx = cmd_mask.index("--apiKeyId")
            if idx + 1 < len(cmd_mask):
                cmd_mask[idx + 1] = "<redacted>"
        if "--apiKey" in cmd_mask:
            idx = cmd_mask.index("--apiKey")
            if idx + 1 < len(cmd_mask):
                cmd_mask[idx + 1] = "<redacted>"

        try:
            # Run in the script's directory to handle relative paths
            # Log the masked command
            LOGGER.info(f"Running vde.py command: {' '.join(cmd_mask)}")
            process = subprocess.run(
                cmd, 
                check=True, 
                cwd=self.export_script_dir,
                capture_output=True,
                text=True
            )
            
            # Log important parts of the output but not everything to avoid spam
            stdout_lines = process.stdout.strip().split('\n')
            if stdout_lines:
                # Log first 5 lines and last 5 lines (if there's a lot of output)
                if len(stdout_lines) > 10:
                    LOGGER.info("vde.py output first lines:")
                    for line in stdout_lines[:5]:
                        LOGGER.info(f"  {line}")
                    LOGGER.info("...")
                    LOGGER.info("vde.py output last lines:")
                    for line in stdout_lines[-5:]:
                        LOGGER.info(f"  {line}")
                else:
                    # Just log all if it's not too much
                    LOGGER.info("vde.py output:")
                    for line in stdout_lines:
                        LOGGER.info(f"  {line}")
            
            if not os.path.exists(output_file):
                raise FileNotFoundError("vde.py ran but raw_export.xlsx was not created.")
            
            LOGGER.info(f"Generated raw data at {output_file}")
            return output_file
        except subprocess.CalledProcessError as e:
            LOGGER.error(f"Failed to run vde.py: {e}")
            if e.stderr:
                stderr_lines = e.stderr.strip().split('\n')
                LOGGER.error("vde.py stderr output:")
                for line in stderr_lines:
                    LOGGER.error(f"  {line}")
            raise RuntimeError(f"vde.py export failed: {e}")

    def _recalculate_with_libreoffice(self, excel_file):
        """Use LibreOffice to ensure formulas are recalculated."""
        if not self.libreoffice_available:
            LOGGER.warning("Skipping LibreOffice recalculation (not available)")
            return
            
        try:
            # Create a temporary directory for the conversion
            with tempfile.TemporaryDirectory() as temp_dir:
                # Get filename without path
                filename = os.path.basename(excel_file)
                
                # Construct LibreOffice command for silent recalculation
                cmd = [
                    "libreoffice", "--headless", "--calc", 
                    "--convert-to", "xlsx", 
                    "--outdir", temp_dir,
                    excel_file
                ]
                
                LOGGER.info(f"Recalculating formulas in {excel_file} with LibreOffice")
                result = subprocess.run(cmd, capture_output=True, text=True, check=False)
                
                if result.returncode != 0:
                    LOGGER.warning(f"LibreOffice recalculation warning: {result.stderr}")
                
                # Get the converted file name - LibreOffice might change the extension
                converted_files = [f for f in os.listdir(temp_dir) if f.startswith(os.path.splitext(filename)[0])]
                
                if converted_files:
                    converted_file = os.path.join(temp_dir, converted_files[0])
                    shutil.copy(converted_file, excel_file)
                    LOGGER.info(f"Recalculated file saved back to {excel_file}")
                else:
                    LOGGER.warning(f"LibreOffice did not create converted file: {os.listdir(temp_dir)}")
        except Exception as e:
            LOGGER.error(f"Error during LibreOffice recalculation: {e}")

    def update_master_workbook(self, raw_file, master_template, target_date=None):
        """
        Copy data from raw export to master workbook and update formulas.
        
        Args:
            raw_file: Path to the raw data export file
            master_template: Path to the master template to use as a base
            target_date: Date for the target workbook (if None, extract from template)
        """
        # Determine the target date for the new workbook
        if target_date is None:
            # Extract date from the template filename
            template_date = self._extract_date_from_filename(master_template)
            if template_date:
                # The new file should be for the day after the template date
                target_date = template_date + timedelta(days=1)
                LOGGER.info(f"Using date from template: {template_date.strftime('%Y-%m-%d')} → {target_date.strftime('%Y-%m-%d')}")
            else:
                # Fall back to current date if we can't extract
                target_date = datetime.now(tz.gettz(self.timezone))
                LOGGER.warning(f"Could not extract date from template filename, using current date: {target_date.strftime('%Y-%m-%d')}")
        
        # Format the date string for the filename
        target_str = target_date.strftime("%m%d%y")
        new_master_file = os.path.join(self.work_dir, f"3895th_{target_str}.xlsx")

        LOGGER.info(f"Creating new master workbook: {new_master_file} for {target_date.strftime('%Y-%m-%d')}")
        shutil.copy(master_template, new_master_file)
        
        try:
            # Use openpyxl for data transfer with keep_vba=True to preserve charts and macros
            raw_wb = openpyxl.load_workbook(raw_file)
            master_wb = openpyxl.load_workbook(new_master_file, data_only=False, keep_vba=True)

            # Check if RAW sheet exists in raw workbook
            if "RAW" not in raw_wb.sheetnames:
                LOGGER.error(f"'RAW' sheet not found in raw export workbook")
                raise ValueError("Raw export workbook missing 'RAW' sheet")
                
            raw_sheet = raw_wb["RAW"]
            
            # Check if "Raw Import" tab exists in master workbook
            if "Raw Import" not in master_wb.sheetnames:
                LOGGER.error(f"'Raw Import' sheet not found in master workbook")
                raise ValueError("Master workbook missing 'Raw Import' sheet")
                
            import_sheet = master_wb["Raw Import"]

            # Clear existing data in the import sheet
            LOGGER.info("Clearing existing data in Raw Import sheet")
            for row in import_sheet.iter_rows():
                for cell in row:
                    cell.value = None

            # Copy data more efficiently using cell references
            LOGGER.info("Copying data from RAW to Raw Import")
            row_count = 0
            for row_idx, row in enumerate(raw_sheet.rows, start=1):
                for col_idx, cell in enumerate(row, start=1):
                    import_sheet.cell(row=row_idx, column=col_idx).value = cell.value
                row_count += 1

            # Save the workbook with updated data
            master_wb.save(new_master_file)
            LOGGER.info(f"Copied {row_count} rows to Raw Import tab")
            
            # Use LibreOffice for formula recalculation if available
            self._recalculate_with_libreoffice(new_master_file)
            
            LOGGER.info(f"Updated workbook saved at {new_master_file}")
            return new_master_file
        except Exception as e:
            LOGGER.error(f"Error updating master workbook: {e}")
            if os.path.exists(new_master_file):
                try:
                    # Try to keep the file for troubleshooting
                    error_file = f"{new_master_file}.error"
                    shutil.copy(new_master_file, error_file)
                    LOGGER.info(f"Saved error state to {error_file}")
                except Exception:
                    pass
            raise

    def process(self, master_template, target_date=None):
        """
        Main processing function: run export, update master workbook.
        
        Args:
            master_template: Path to the master template to use
            target_date: Specific date to process (if None, use date from template)
        """
        os.makedirs(self.work_dir, exist_ok=True)
        LOGGER.info(f"Starting workbook processing with template {master_template}")
        
        # Determine the date to process
        if target_date is None:
            # If no target date provided, extract from master template
            template_date = self._extract_date_from_filename(master_template)
            if template_date:
                # Use the day after the template date as the target date
                target_date = template_date + timedelta(days=1)
                LOGGER.info(f"Using date from template filename: processing data for {target_date.strftime('%Y-%m-%d')}")
            else:
                # Default to yesterday if we can't extract from filename
                target_date = self.get_yesterday_date()
                LOGGER.warning(f"Could not extract date from template, defaulting to yesterday: {target_date.strftime('%Y-%m-%d')}")
        else:
            LOGGER.info(f"Using provided target date: {target_date.strftime('%Y-%m-%d')}")
        
        raw_file = os.path.join(self.work_dir, "raw_export.xlsx")
        
        # Pass the target date to run_vde_export to use the correct date for data
        self.run_vde_export(raw_file, target_date)
        
        # Pass both the master template and target date to update_master_workbook
        return self.update_master_workbook(raw_file, master_template, target_date)